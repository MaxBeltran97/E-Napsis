import flask
from flask import request, Flask
from models.calendarCourseAttendance import *
from models.participant import *
from strgen import StringGenerator
import os
import openpyxl


app = Flask(__name__)

UPLOAD_FOLDER = 'assets/excelAttendance'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

calendarAttendance = flask.Blueprint('calendarAttendance', __name__)



@calendarAttendance.route('/api/calendar/attendance', methods=['POST'])
def add_attendance():
    try:

        date = request.json['date']
        calendar_id = request.json['calendarCourse_id']

        new_attendance = CalendarCourseAttendance(calendar_id, date, '-')

        attendance_id = CalendarCourseAttendance.query.filter_by(calendarCourse_id = calendar_id)


        if(bool(attendance_id.first()) == False):
            participants = Participant.query.filter_by(calendarCourse_id = calendar_id)
            wb = openpyxl.Workbook()
            ws = wb.active
            cont = 2
            for i in participants:
                ws[f'A{cont}'] = i._id
                cont += 1


            db.session.add(new_attendance)
            db.session.commit()

            ws['B1'] = new_attendance._id
            randomName = StringGenerator(
                    "[\l\d]{20}").render_list(1, unique=True)[0]
            nuevoNombreFile = randomName + ".xlsx" 
            

            path = os.path.join(
                app.config["UPLOAD_FOLDER"],nuevoNombreFile
            )
            wb.save(path)
            new_attendance.excelPath = nuevoNombreFile
            db.session.commit()
        
        else:
            excelPath = ''
            excelPathComplete = ''
            db.session.add(new_attendance)
            db.session.commit()
            for i in attendance_id:
                excelPath = i.excelPath
                excelPathComplete = os.path.join(app.config["UPLOAD_FOLDER"], i.excelPath)
                break

            wb = openpyxl.load_workbook(excelPathComplete)
            ws = wb.active
            ws.cell(row=1, column=ws.max_column + 1).value = new_attendance._id
            wb.save(excelPathComplete)

            new_attendance.excelPath = excelPath
            db.session.commit()

        return {
                "ok": True
            }, 201
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al añadir la asistencia de un curso"
        }, 500
    finally:
        db.session.close()

@calendarAttendance.route('/api/calendar/attendance/<_idCalendar>', methods=['GET'])
def get_attendances(_idCalendar):
    try:
        all_attendances = CalendarCourseAttendance.query.filter_by(calendarCourse_id = _idCalendar)
        result = calendar_course_attendance_schemas.dump(all_attendances)

        return {
            "ok": True,
            "attendances": result
        }, 200
    except Exception as e:
        print(e)
        return {
            'ok': False,
            'msg': 'Error al obtener las asistencias'
        }

@calendarAttendance.route('/api/calendar/attendance/attendance/<_idAttendance>', methods=['GET'])
def get_attendance_id(_idAttendance):
    try:
        attendance = CalendarCourseAttendance.query.get(_idAttendance)

        return {
            "ok": True,
            "attendance": attendance.serialize()
        }, 200
    except Exception as e:
        print(e)
        return {
            'ok': False,
            'msg': 'Error al obtener la asistencia'
        }

@calendarAttendance.route('/api/calendar/attendance/<_id>', methods=['PUT'])
def update_attendance(_id):

    try:

        attendance = CalendarCourseAttendance.query.get(_id)

        date = request.json['date']

        attendance.date = date

        db.session.commit()

        return {
            "ok": True,
            "asistencia": attendance.serialize()
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al actualizar la asistencia"
        }, 500

@calendarAttendance.route('/api/calendar/attendance/<_id>', methods=['DELETE'])
def delete_attendance(_id):
    try:

        attendance = CalendarCourseAttendance.query.get(_id)

        path = os.path.join(
            app.config["UPLOAD_FOLDER"], attendance.excelPath
        )
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for cell in ws[1]:
            if cell.value == attendance._id:
                ws.delete_cols(cell.column)
                break

        wb.save(path)

        db.session.delete(attendance)
        db.session.commit()

        return {
            "ok": True,
            "asistencia": attendance.serialize()
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al eliminar la asistencia"
        }, 500

@calendarAttendance.route('/api/calendar/attendance/upload/<_id>', methods=['PUT'])
def upload_attendance(_id):

    try:
        
        attendance = CalendarCourseAttendance.query.get(_id)
        
        attendances = request.json["attendances"]
        columna = ''

        path = os.path.join(
            app.config["UPLOAD_FOLDER"],attendance.excelPath
            )

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        for cell in ws[1]:
            if cell.value == attendance._id:
                columna = cell.column
                break
        

        for p in attendances:
            fila = ''
            for i in range(1, ws.max_row + 1):
                if(ws[i][0].value == p['participant_id']):
                    fila = i
            ws[fila][columna-1].value = int(p['attendance'])
 
        wb.save(path)

        return {
            "ok": True
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al subir la asistencia"
        }, 500

@calendarAttendance.route('/api/calendar/attendance/file/<_id>', methods=['GET'])
def get_attendance(_id):

    try:
        
        attendances = []
        attendance = CalendarCourseAttendance.query.get(_id)

        columna = ''

        path = os.path.join(
            app.config["UPLOAD_FOLDER"],attendance.excelPath
            )

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        for cell in ws[1]:
            if cell.value == attendance._id:
                columna = cell.column
                break
        
        for i in range(2, ws.max_row + 1):

            attendance = None
            if(ws[i][columna-1].value != None):
                attendance = str(ws[i][columna-1].value)
            else:
                attendance = ''

            attendances.append(
                {
                    "participant_id": ws[i][0].value,
                    "attendance": attendance
                }
            )

        wb.save(path)

        return {
            "ok": True,
            "attendances": attendances
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al obtener la asistencia"
        }, 500