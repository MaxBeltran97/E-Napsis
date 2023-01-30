import flask
from flask import request, Flask
from models.calendarCourseEvaluation import *
from models.participant import *
import pandas as pd
from strgen import StringGenerator
import os
import openpyxl


app = Flask(__name__)

UPLOAD_FOLDER = 'assets/excelEvaluation'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

evaluation = flask.Blueprint('evaluation', __name__)


@evaluation.route('/api/calendar/evaluation', methods=['POST'])
def add_evaluation():
    try:
        evaluationDate = request.json['evaluationDate']
        percentage = request.json['percentage']
        title = request.json['title']
        calendar_id = request.json['calendarCourse_id']

        new_evaluation = CalendarCourseEvaluation(calendar_id, evaluationDate, percentage, title, '-')

        evaluation_id = CalendarCourseEvaluation.query.filter_by(calendarCourse_id = calendar_id)
       
        
        if(bool(evaluation_id.first()) == False):
            participants = Participant.query.filter_by(calendarCourse_id = calendar_id)
            wb = openpyxl.Workbook()
            ws = wb.active
            cont = 2
            for i in participants:
                ws[f'A{cont}'] = i._id
                cont += 1
                

            db.session.add(new_evaluation)
            db.session.commit()
            
            ws['B1'] = new_evaluation._id
            randomName = StringGenerator(
                    "[\l\d]{20}").render_list(1, unique=True)[0]
            nuevoNombreFile = randomName + ".xlsx" 
            

            path = os.path.join(
                app.config["UPLOAD_FOLDER"],nuevoNombreFile
            )
            wb.save(path)
            new_evaluation.excelPath = nuevoNombreFile
            db.session.commit()
            
        else:
            excelPath = ''
            excelPathComplete = ''
            db.session.add(new_evaluation)
            db.session.commit()
            for i in evaluation_id:
                excelPath = i.excelPath
                excelPathComplete = os.path.join(app.config["UPLOAD_FOLDER"], i.excelPath)    
                break
            
            wb = openpyxl.load_workbook(excelPathComplete)
            ws = wb.active
            ws.cell(row=1, column=ws.max_column + 1).value = new_evaluation._id
            wb.save(excelPathComplete)
            
            new_evaluation.excelPath = excelPath
            db.session.commit()

        
        return {
                "ok": True
            }, 201
    except Exception as e:
        print('aa')
        print(e)
        return {
            "ok": False,
            "msg": "Error al calendarizar un curso"
        }, 500
    finally:
        db.session.close()

@evaluation.route('/api/calendar/evaluation/<_idCalendar>', methods=['GET'])
def get_evaluations(_idCalendar):
    try:
        all_evaluations = CalendarCourseEvaluation.query.filter_by(calendarCourse_id = _idCalendar)
        result = calendarCourseEvaluations_schemas.dump(all_evaluations)

        return {
            "ok": True,
            "evaluations": result
        }, 200
    except Exception as e:
        print(e)
        return {
            'ok': False,
            'msg': 'Error al obtener las evaluaciones'
        }

@evaluation.route('/api/calendar/evaluation/evaluation/<_idEvaluation>', methods=['GET'])
def get_evaluation(_idEvaluation):
    try:
        evaluation = CalendarCourseEvaluation.query.get(_idEvaluation)

        return {
            "ok": True,
            "evaluation": evaluation.serialize()
        }, 200
    except Exception as e:
        print(e)
        return {
            'ok': False,
            'msg': 'Error al obtener la evaluacion'
        }

@evaluation.route('/api/calendar/evaluation/<_id>', methods=['PUT'])
def update_evaluation(_id):
    try:
        
        evaluation = CalendarCourseEvaluation.query.get(_id)
        
        evaluationDate = request.json['evaluationDate']
        percentage = request.json['percentage']
        title = request.json['title']
        
        evaluation.evaluationDate = evaluationDate
        evaluation.percentage = percentage
        evaluation.title = title

        db.session.commit()
        
        return {
            "ok": True,
            "evaluation": evaluation.serialize()
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al actualizar la evaluación"
        }, 500

@evaluation.route('/api/calendar/evaluation/<_id>', methods=['DELETE'])
def delete_evaluation(_id):
    try:
        
        evaluation = CalendarCourseEvaluation.query.get(_id)
        
        path = os.path.join(
                app.config["UPLOAD_FOLDER"],evaluation.excelPath
            )
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for cell in ws[1]:
            if cell.value == evaluation._id:
                ws.delete_cols(cell.column)
                break

        wb.save(path)

        db.session.delete(evaluation)
        db.session.commit()
        
        return {
            "ok": True,
            "evaluation": evaluation.serialize()
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al eliminar la evaluación"
        }, 500

@evaluation.route('/api/calendar/evaluation/grades/<_id>', methods=['PUT'])
def upload_grades(_id):
    try:
        
        evaluation = CalendarCourseEvaluation.query.get(_id)

        grades = request.json["grades"]
        columna = ''

        path = os.path.join(
                app.config["UPLOAD_FOLDER"],evaluation.excelPath
            )
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        
        for cell in ws[1]:
            if cell.value == evaluation._id:
                columna = cell.column
                break

        for p in grades:
            fila = ''
            for i in range(1, ws.max_row + 1):
                if(ws[i][0].value == p['participant_id']):
                    fila = i
            ws[fila][columna-1].value = float(p['grade'].replace(",", "."))

        wb.save(path)

        return {
            "ok": True
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al actualizar las notas"
        }, 500

@evaluation.route('/api/calendar/evaluation/grades/<_id>', methods=['GET'])
def get_grades(_id):
    try:
        
        grades = []
        evaluation = CalendarCourseEvaluation.query.get(_id)

        columna = ''

        path = os.path.join(
                app.config["UPLOAD_FOLDER"],evaluation.excelPath
            )
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        
        for cell in ws[1]:
            if cell.value == evaluation._id:
                columna = cell.column
                break

        for i in range(2, ws.max_row + 1):

            grade = None
            if(ws[i][columna-1].value != None):
                grade = (str(ws[i][columna-1].value)).replace(".", ",")

            grades.append(
                {
                    "participant_id": ws[i][0].value, 
                    "grade": grade
                }
            )

        wb.save(path)

        return {
            "ok": True,
            "grades": grades
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al actualizar las notas"
        }, 500