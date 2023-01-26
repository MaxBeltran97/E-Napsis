import flask
from flask import request, Flask
from models.participant import *
from models.calendarCourseEvaluation import *
import os
import openpyxl
import pandas as pd




app = Flask(__name__)

UPLOAD_FOLDER = 'assets/excelEvaluation'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

participants = flask.Blueprint('participants', __name__)



@participants.route('/api/participant', methods=['POST'])
def add_participant():
    try:
        calendarCourse_id = request.json['calendarCourse_id']
        participantType = request.json['participantType']
        company_id = request.json['company_id']
        nationalityType = request.json['nationalityType']
        rut = request.json['rut']
        fullName = request.json['fullName']
        lastName = request.json['lastName']
        motherLastName = request.json['motherLastName']
        institution = request.json['institution']
        email = request.json['email']
        gender = request.json['gender']
        position = request.json['position']

        new_participant = Participant(calendarCourse_id, participantType, company_id, nationalityType,
                                           rut, fullName, lastName, motherLastName, institution, email, gender, position)

        db.session.add(new_participant)
        db.session.commit()

        evaluation_id = CalendarCourseEvaluation.query.filter_by(calendarCourse_id = calendarCourse_id)

        if(bool(evaluation_id.first()) == True):
            
            excelPathComplete = ''
            for i in evaluation_id:
                
                excelPathComplete = os.path.join(app.config["UPLOAD_FOLDER"], i.excelPath)    
                break
            wb = openpyxl.load_workbook(excelPathComplete)
            ws = wb.active
            ws.cell(row=ws.max_row + 1, column=1).value = new_participant._id
            wb.save(excelPathComplete)

        return {
            "ok": True,
            "participant": new_participant.serialize()
        }, 201
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al guardar el participante"
        }, 500
    finally:
        db.session.close()


@participants.route('/api/participant', methods=['GET'])
def get_participants():
    try:
        all_participants = Participant.query.all()
        result = participants_schema.dump(all_participants)
        return {
            "ok": True,
            "participants": result
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al obtener los participantes"
        }, 500


@participants.route('/api/participant/<_id>', methods=['GET'])
def get_participant(_id):
    try:
        participant = Participant.query.get(_id)
        return {
            "ok": True,
            "participant": participant.serialize()
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al obtener el participante"
        }, 500


@participants.route('/api/participant/<_id>', methods=['PUT'])
def update_participant(_id):
    try:
        participant = Participant.query.get(_id)

        calendarCourse_id = request.json['calendarCourse_id']
        participantType = request.json['participantType']
        company_id = request.json['company_id']
        nationalityType = request.json['nationalityType']
        rut = request.json['rut']
        fullName = request.json['fullName']
        lastName = request.json['lastName']
        motherLastName = request.json['motherLastName']
        institution = request.json['institution']
        email = request.json['email']
        gender = request.json['gender']
        position = request.json['position']

        isChangeCourse = False
        
        courseBefore = ''

        if(calendarCourse_id != participant.calendarCourse_id):
            courseBefore = participant.calendarCourse_id
            isChangeCourse = True

        participant.calendarCourse_id = calendarCourse_id
        participant.participantType = participantType
        participant.company_id = company_id
        participant.nationalityType = nationalityType
        participant.rut = rut
        participant.fullName = fullName
        participant.lastName = lastName
        participant.motherLastName = motherLastName
        participant.institution = institution
        participant.email = email
        participant.gender = gender
        participant.position = position

        db.session.commit()

        
        if(isChangeCourse == True):
            evaluation_id = CalendarCourseEvaluation.query.filter_by(calendarCourse_id =courseBefore)

            if(bool(evaluation_id.first()) == True):
            
                excelPathComplete = ''
                for i in evaluation_id:
                    excelPathComplete = os.path.join(app.config["UPLOAD_FOLDER"], i.excelPath)    
                    break

                wb = openpyxl.load_workbook(excelPathComplete)
                ws = wb.active

                for i in range(1, ws.max_row + 1):
                    if(ws[i][0].value == participant._id):
                        ws.delete_rows(i)
                        break
                wb.save(excelPathComplete)

            new_course = CalendarCourseEvaluation.query.filter_by(calendarCourse_id = participant.calendarCourse_id)

            if(bool(new_course.first()) == True):
                
                excelPathComplete = ''
                for i in new_course:
                    
                    excelPathComplete = os.path.join(app.config["UPLOAD_FOLDER"], i.excelPath)    
                    break
                wb = openpyxl.load_workbook(excelPathComplete)
                ws = wb.active
                ws.cell(row=ws.max_row + 1, column=1).value = participant._id
                wb.save(excelPathComplete)

        return {
            "ok": True,
            "participant": participant.serialize()
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al actualizar el participante"
        }, 500
    finally:
        db.session.close()


@participants.route('/api/participant/<_id>', methods=['DELETE'])
def delete_participant(_id):
    try:
        participant = Participant.query.get(_id)

        db.session.delete(participant)
        db.session.commit()


        evaluation_id = CalendarCourseEvaluation.query.filter_by(calendarCourse_id = participant.calendarCourse_id)


        if(bool(evaluation_id.first()) == True):
            
            excelPathComplete = ''
            for i in evaluation_id:
                excelPathComplete = os.path.join(app.config["UPLOAD_FOLDER"], i.excelPath)    
                break
            
            wb = openpyxl.load_workbook(excelPathComplete)
            ws = wb.active

            for i in range(1, ws.max_row + 1):
                if(ws[i][0].value == participant._id):
                    ws.delete_rows(i)
                    break
            wb.save(excelPathComplete)

        return {
            "ok": True,
            "participant": participant.serialize()
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al eliminar el participante"
        }, 500
    finally:
        db.session.close()


@participants.route('/api/participant/uploadfile', methods=['POST'])
def upload_file_participants():
    try:
        print('entre')
        calendarCourse_id = request.form['calendarCourse_id']
        company_id = request.form['company_id']
        file = request.files["excel"]

        msg = "posiciones duplicadas: "
        df = pd.read_excel(file)

        for i in range(len(df.index)):
            nationalityType = df.loc[i, 'NACIONALIDAD']
            participantType = df.loc[i, 'CARGO DESEMPEÑADO']
            rut = df.loc[i, 'RUT']
            fullName = df.loc[i, 'NOMBRE']
            lastName = df.loc[i, 'PATERNO']
            motherLastName = df.loc[i, 'MATERNO']
            institution = df.loc[i, 'ESTABLECIMIENTO']
            email = df.loc[i, 'EMAIL']
            gender = df.loc[i, 'GENERO']
            position = df.loc[i, 'POSICION']

            try:
                new_participant = Participant(calendarCourse_id, participantType, company_id, nationalityType,
                                                   rut, fullName, lastName, motherLastName, institution, email, gender, position)
                db.session.add(new_participant)
                db.session.commit()
            except Exception as e:
                msg = (msg + f"{i}, ")
                db.session.rollback()
            finally:
                db.session.close()

        return {
            "ok": True,
            "msg": msg
        }, 200
    except Exception as e:
        print(e)
        return {
            "ok": False,
            "msg": "Error al subir el archivo"
        }, 500
